# Demo Excel file
from openpyxl import Workbook

wb = Workbook()

# Create a worksheet named HCMUE
ws = wb.create_sheet("HCMUE", 1)

ws["A1"] = "FIT-HCMUE" 		# Write data in cell "A1"

ws.append([2, 3, 4])			# Insert new rows

wb.save("DemoOpenpyxl.xlsx")		
